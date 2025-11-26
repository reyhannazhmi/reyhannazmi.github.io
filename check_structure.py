import os
import openpyxl

def check_structure():
    print("=== CHECKING PROJECT STRUCTURE ===")
    print(f"Current directory: {os.getcwd()}")
    print(f"Files in current directory: {os.listdir('.')}")
    
    # Cek folder BeeTheOne
    if os.path.exists('BeeTheOne'):
        print(f"✓ BeeTheOne folder exists")
        print(f"Files in BeeTheOne: {os.listdir('BeeTheOne')}")
    else:
        print(f"✗ BeeTheOne folder does not exist!")
        # Coba buat folder
        os.makedirs('BeeTheOne', exist_ok=True)
        print("✓ Created BeeTheOne folder")
    
    # Cek file daftarsaldo.xlsx
    daftarsaldo_path = 'BeeTheOne/daftarsaldo.xlsx'
    if os.path.exists(daftarsaldo_path):
        print(f"✓ daftarsaldo.xlsx exists")
        try:
            wb = openpyxl.load_workbook(daftarsaldo_path)
            print(f"Sheets in daftarsaldo: {wb.sheetnames}")
        except Exception as e:
            print(f"✗ Error reading daftarsaldo: {e}")
    else:
        print(f"✗ daftarsaldo.xlsx does not exist")
    
    # Cek file databasesia.xlsx
    databasesia_path = 'BeeTheOne/databasesia.xlsx'
    if os.path.exists(databasesia_path):
        print(f"✓ databasesia.xlsx exists")
        try:
            wb = openpyxl.load_workbook(databasesia_path)
            print(f"Sheets in databasesia: {wb.sheetnames}")
        except Exception as e:
            print(f"✗ Error reading databasesia: {e}")
    else:
        print(f"✗ databasesia.xlsx does not exist")

if __name__ == "__main__":
    check_structure()