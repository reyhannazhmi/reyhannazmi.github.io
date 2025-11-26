from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, get_flashed_messages
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import os
import openpyxl
import pandas as pd
from functools import wraps
import logging
from markupsafe import Markup
from datetime import datetime

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

app = Flask(__name__)
app.secret_key = 'your_secret_key_here'

# Database configuration
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'BeeTheOne', 'users.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

DATA_DIR = basedir
INVENTORY_FILE = os.path.join(DATA_DIR, 'databasesia.xlsx')
JOURNAL_FILE = os.path.join(DATA_DIR, 'jurnal.xlsx')
SALDO_FILE = os.path.join(DATA_DIR, 'daftarsaldo.xlsx')

MONTH_NAME_TO_NUM = {
    'Januari': '01',
    'Februari': '02',
    'Maret': '03',
    'April': '04',
    'Mei': '05',
    'Juni': '06',
    'Juli': '07',
    'Agustus': '08',
    'September': '09',
    'Oktober': '10',
    'November': '11',
    'Desember': '12',
}


def _is_future_period(tahun_str, bulan_name):
    """Return True if the selected tahun/bulan is in the future compared to today.

    Used to prevent showing Saldo Awal / Neraca Saldo for future periods
    when the previous month has not finished yet.
    """
    try:
        year = int(tahun_str)
    except (ValueError, TypeError):
        return False

    month_code = MONTH_NAME_TO_NUM.get(bulan_name, None)
    if not month_code:
        return False

    try:
        selected_start = datetime(year, int(month_code), 1).date()
    except ValueError:
        return False

    today = datetime.today().date()
    return selected_start > today


def _normalize_excel_date(value):
    """Return date value as YYYY-MM-DD string for comparison."""
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if hasattr(value, 'to_pydatetime'):
        return value.to_pydatetime().strftime('%Y-%m-%d')
    if value is None:
        return ''
    try:
        return str(value).split(' ')[0]
    except Exception:
        return str(value)


def journal_row_exists(ws, tanggal, keterangan, akun, debit, kredit):
    """Check whether a journal row already exists to prevent duplicates."""
    try:
        target_date = tanggal
        if isinstance(tanggal, datetime):
            target_date = tanggal.strftime('%Y-%m-%d')
    except Exception:
        target_date = str(tanggal)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        existing_date = _normalize_excel_date(row[0])
        if existing_date != target_date:
            continue
        existing_keterangan = (row[1] or '').strip()
        existing_akun = (row[2] or '').strip()
        if existing_keterangan != (keterangan or '').strip():
            continue
        if existing_akun != (akun or '').strip():
            continue

        existing_debit = float(row[3] or 0.0)
        existing_kredit = float(row[4] or 0.0)
        if abs(existing_debit - float(debit or 0.0)) < 0.01 and abs(existing_kredit - float(kredit or 0.0)) < 0.01:
            return True
    return False

# Set up logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# User model
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    hashed_password = db.Column(db.String(200), nullable=False)

def safe_float(value):
    """Convert value to float safely"""
    try:
        if value is None:
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def safe_int(value):
    """Convert value to int safely"""
    try:
        if value is None:
            return 0
        return int(value)
    except (ValueError, TypeError):
        return 0


def parse_amount(value):
    """Parse Rupiah input from form.

    Accepts strings like "1000000", "1.000.000", or "1,000,000.50".
    Dots are treated as thousand separators, comma as decimal separator.
    Returns float amount in Rupiah.
    """
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0
    # First remove thousand separators, then normalize decimal separator
    s = s.replace('.', '')
    s = s.replace(',', '.')
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0

def format_rupiah(amount):
    """Format number to Rupiah currency format"""
    try:
        if amount == 0 or amount is None:
            return "Rp 0"
        formatted = f"Rp {float(amount):,.0f}".replace(',', '.')
        return formatted
    except (ValueError, TypeError):
        return "Rp 0"

def format_rupiah_for_report(amount):
    """Format khusus untuk laporan keuangan"""
    try:
        if amount is None or amount == 0:
            return "Rp 0"
        result = f"Rp {int(amount):,}".replace(',', '.')
        return result
    except Exception:
        return "Rp 0"

# Template filter
@app.template_filter('safe_currency')
def safe_currency_filter(value):
    try:
        if value is None:
            return ""
        val = float(value)
        return f"Rp {val:,.0f}".replace(",", ".")
    except (ValueError, TypeError):
        return ""

def load_inventory():
    """Membaca data inventory dari file Excel dengan struktur yang benar"""
    try:
        # Baca file Excel
        excel_file = INVENTORY_FILE
        df = pd.read_excel(excel_file, sheet_name='Inventory')

        inventory_data = []
        default_code_index = 1
        assigned_codes = set()

        # Iterasi melalui setiap baris di sheet Inventory
        for index, row in df.iterrows():
            raw_item_code = row.iloc[0]
            if raw_item_code is None or str(raw_item_code).strip() == '' or str(raw_item_code).strip().lower() == 'no item':
                # Assign default item_code in the format ITEM-XXX
                while True:
                    default_code = f'ITEM-{default_code_index:03d}'
                    default_code_index += 1
                    if default_code not in assigned_codes:
                        item_code = default_code
                        assigned_codes.add(default_code)
                        break
                logger.warning(f"Row {index}: Invalid or missing item_code '{raw_item_code}', assigned default code '{item_code}'.")
            else:
                # Normalize item_code: strip spaces and convert to uppercase
                item_code = str(raw_item_code).strip().upper()
                assigned_codes.add(item_code)

            stock = int(row['Stock Remaining']) if not pd.isna(row['Stock Remaining']) else 0
            cost_price_unit = float(row['Price']) if not pd.isna(row['Price']) else 0.0

            # Ambil harga jual per unit dari kolom yang benar
            selling_price_unit = 0.0
            try:
                if 'Harga Jual' in df.columns and not pd.isna(row['Harga Jual']):
                    selling_price_unit = float(row['Harga Jual'])
                elif 'Unnamed: 8' in df.columns and not pd.isna(row['Unnamed: 8']):
                    # Di file saat ini, kolom terakhir (Unnamed: 8) berisi harga jual per unit
                    selling_price_unit = float(row['Unnamed: 8'])
            except Exception as e:
                logger.warning(f"Error parsing selling price for row {index}: {e}")
                selling_price_unit = 0.0

            cost_total = cost_price_unit * stock
            selling_total = selling_price_unit * stock
            gross_profit_total = selling_total - cost_total
            is_stock_flag = stock > 0

            item_data = {
                'item_code': item_code,
                'name': str(row.iloc[1]) if not pd.isna(row.iloc[1]) else 'Unknown Product',
                'stock': stock,
                'cost_price': cost_price_unit,
                'selling_price': selling_price_unit,
                'gross_profit': gross_profit_total,
                'is_stock': is_stock_flag,
                'cost_price_stock': cost_total,
                'selling_price_stock': selling_total,
                'selling_price_total': selling_total,
                'cost_price_total': cost_total
            }
            inventory_data.append(item_data)

        logger.debug(f"Loaded {len(inventory_data)} inventory items from Excel")
        return inventory_data

    except FileNotFoundError as fnfe:
        logger.error(f"Excel file 'databasesia.xlsx' not found: {fnfe}")
        # Fallback data
        fallback = [
            {'item_code': 'ITEM-001', 'name': 'Madu Multiflora', 'stock': 34, 'cost_price': 84000, 'selling_price': 105000, 'gross_profit': (105000-84000)*34, 'is_stock': False},
            {'item_code': 'ITEM-002', 'name': 'Madu Klengkeng', 'stock': 19, 'cost_price': 100000, 'selling_price': 125000, 'gross_profit': (125000-100000)*19, 'is_stock': False},
            {'item_code': 'ITEM-003', 'name': 'Kapuk Randu', 'stock': 22, 'cost_price': 96000, 'selling_price': 120000, 'gross_profit': (120000-96000)*22, 'is_stock': False},
        ]
        return fallback
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return []

# Routes
@app.route('/')
def home():
    return render_template('index.html')

@app.route('/bees')
def bees():
    return render_template('bees.html')

@app.route('/inventory')
@login_required
def inventory():
    inventory_data = load_inventory()
    total_cost_price_stock = sum(item['cost_price'] * item['stock'] for item in inventory_data)
    logger.debug(f"Inventory data loaded with {len(inventory_data)} items.")
    return render_template('inventory.html', inventory_data=inventory_data, total_cost_price_stock=total_cost_price_stock)

@app.route('/dashboard')
@login_required
def dashboard():
    """Route untuk dashboard dengan data real dari Excel"""
    inventory_data = load_inventory()
    
    total_inventory_value = sum(item['cost_price'] * item['stock'] for item in inventory_data)
    total_products = len(inventory_data)
    total_gross_profit = sum((item['selling_price'] - item['cost_price']) * item['stock'] for item in inventory_data)
    
    low_stock_items = [item for item in inventory_data if item['stock'] < 10]
    low_stock_count = len(low_stock_items)
    
    try:
        wb = openpyxl.load_workbook(JOURNAL_FILE)
        ws = wb['Journal']
        journal_entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and len(row) >= 5:
                journal_entries.append({
                    'tanggal': row[0].strftime('%Y-%m-%d') if hasattr(row[0], 'strftime') else str(row[0]),
                    'keterangan': row[1] if row[1] else '',
                    'akun': row[2] if row[2] else '',
                    'debit': float(row[3]) if row[3] else 0,
                    'kredit': float(row[4]) if row[4] else 0
                })
        journal_entries = journal_entries[-5:]
    except Exception as e:
        logger.warning(f"Error loading journal entries: {e}")
        journal_entries = [
            {'tanggal': '2024-01-15', 'keterangan': 'Penjualan Madu Multiflora', 'akun': 'Penjualan', 'debit': 105000, 'kredit': 0},
            {'tanggal': '2024-01-15', 'keterangan': 'Penjualan Madu Multiflora', 'akun': 'Kas', 'debit': 0, 'kredit': 105000},
            {'tanggal': '2024-01-14', 'keterangan': 'Pembelian Botol', 'akun': 'Persediaan', 'debit': 3450000, 'kredit': 0},
        ]
    
    recent_activities = [
        {'tanggal': '2024-01-15', 'produk': 'Madu Multiflora', 'jenis': 'Penambahan', 'qty': 50, 'keterangan': 'Restock dari supplier'},
        {'tanggal': '2024-01-14', 'produk': 'Madu Klengkeng', 'jenis': 'Pengurangan', 'qty': 7, 'keterangan': 'Penjualan'},
        {'tanggal': '2024-01-13', 'produk': 'Madu Hutan', 'jenis': 'Penambahan', 'qty': 20, 'keterangan': 'Restock'}
    ]
    
    return render_template('dashboard.html', 
                         journal_entries=journal_entries,
                         recent_activities=recent_activities,
                         total_inventory_value=total_inventory_value,
                         total_products=total_products,
                         total_gross_profit=total_gross_profit,
                         low_stock_items=low_stock_items,
                         low_stock_count=low_stock_count)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            error = "Username and password must be provided."
            return render_template('register.html', error=error)
        # Check if user exists
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            error = "Username already exists. Please choose a different username."
            return render_template('register.html', error=error)
        # Create user with hashed password
        hashed_pw = generate_password_hash(password)
        new_user = User(username=username, hashed_password=hashed_pw)
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        if not username or not password:
            return render_template('login.html', error='Please enter username and password')
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.hashed_password, password):
            session['user'] = username
            return redirect(url_for('stock_card'))
        else:
            return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')

@app.route('/stock_card')
@login_required
def stock_card():
    inventory_data = load_inventory()
    selected_product = request.args.get('product')
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')
    stock_card_data = []
    item_code = ''

    try:
        if selected_product and tahun == '2025' and bulan == 'November':

            # ─────────────────────────────────────────────────────────────
            # Ambil harga modal & saldo awal dari inventory
            # ─────────────────────────────────────────────────────────────
            cost_price = 0
            initial_qty = 0

            for item in inventory_data:
                if item['name'] == selected_product:
                    cost_price = item['cost_price']
                    initial_qty = safe_int(item.get('stock', 0) or 0)
                    item_code = item['item_code']
                    break

            # ─────────────────────────────────────────────────────────────
            # Load jurnal
            # ─────────────────────────────────────────────────────────────
            try:
                wb = openpyxl.load_workbook(JOURNAL_FILE)
                ws = wb['Journal']
                journal_entries = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] and len(row) >= 5:
                        journal_entries.append({
                            'tanggal': row[0],
                            'keterangan': row[1],
                            'akun': row[2],
                            'debit': row[3] if row[3] else 0,
                            'kredit': row[4] if row[4] else 0
                        })
            except:
                journal_entries = []

            # ─────────────────────────────────────────────────────────────
            # Saldo awal
            # ─────────────────────────────────────────────────────────────
            balance_qty = initial_qty
            balance_price = cost_price
            balance_total = balance_qty * balance_price

            stock_card_data.append({
                'date': 'Saldo Awal',
                'description': 'Saldo awal persediaan',
                'in_qty': initial_qty,
                'in_price': format_rupiah(cost_price),
                'in_total': format_rupiah(initial_qty * cost_price),
                'out_qty': None,
                'out_price': None,
                'out_total': None,
                'balance_qty': balance_qty,
                'balance_price': format_rupiah(balance_price),
                'balance_total': format_rupiah(balance_total)
            })

            # Daftar akun persediaan yang valid
            valid_inventory_accounts = [
                "1-1300",
                "Persediaan Barang Dagang",
                "Persediaan Madu"
            ]

            # ─────────────────────────────────────────────────────────────
            # Proses jurnal
            # ─────────────────────────────────────────────────────────────
            for entry in journal_entries:

                # Filter berdasarkan tahun & bulan
                entry_date = str(entry['tanggal']).split('-')
                if len(entry_date) < 2:
                    continue
                entry_year = entry_date[0]
                entry_month = entry_date[1]

                month_names = {
                    '01': 'Januari', '02': 'Februari', '03': 'Maret', '04': 'April', '05': 'Mei', '06': 'Juni',
                    '07': 'Juli', '08': 'Agustus', '09': 'September', '10': 'Oktober', '11': 'November', '12': 'Desember'
                }

                if entry_year != tahun or month_names.get(entry_month, '') != bulan:
                    continue

                # ─────────────────────────────────────────────────────────────
                # Filter: hanya akun persediaan yang boleh mempengaruhi stok
                # ─────────────────────────────────────────────────────────────
                akun = entry['akun'].lower()

                is_inventory_related = any(
                    valid.lower() in akun
                    for valid in valid_inventory_accounts
                )

                if not is_inventory_related:
                    continue

                # Add filter: check if the journal entry keterangan contains selected product name (case-insensitive)
                if selected_product and selected_product.lower() not in (entry['keterangan'] or '').lower():
                    continue

                # ─────────────────────────────────────────────────────────────
                # Hitung transaksi in/out
                # ─────────────────────────────────────────────────────────────
                in_qty = out_qty = None
                in_price = out_price = None
                in_total = out_total = None

                # Debit = masuk (pembelian)
                if entry['debit'] > 0:
                    in_total = entry['debit']
                    in_price = cost_price
                    in_qty = in_total / cost_price if cost_price > 0 else 0

                    balance_qty += in_qty
                    balance_total += in_total
                    balance_price = balance_total / balance_qty if balance_qty > 0 else 0

                # Kredit = keluar (penjualan/COGS)
                elif entry['kredit'] > 0:
                    out_total = entry['kredit']
                    out_price = cost_price
                    out_qty = out_total / cost_price if cost_price > 0 else 0

                    balance_qty -= out_qty
                    balance_total -= out_total
                    balance_price = balance_total / balance_qty if balance_qty > 0 else 0

                stock_card_data.append({
                    'date': entry['tanggal'],
                    'description': entry['keterangan'],
                    'in_qty': in_qty,
                    'in_price': format_rupiah(in_price) if in_price else '',
                    'in_total': format_rupiah(in_total) if in_total else '',
                    'out_qty': out_qty,
                    'out_price': format_rupiah(out_price) if out_price else '',
                    'out_total': format_rupiah(out_total) if out_total else '',
                    'balance_qty': balance_qty,
                    'balance_price': format_rupiah(balance_price),
                    'balance_total': format_rupiah(balance_total)
                })

    except Exception as e:
        logger.error(f"Error in stock_card route: {e}")
        stock_card_data = []

    return render_template(
        'stock_card.html',
        inventory_data=inventory_data,
        stock_card_data=stock_card_data,
        selected_product=selected_product,
        tahun=tahun,
        bulan=bulan,
        item_code=item_code
    )

@app.route('/journal')
@login_required
def journal():
    try:
        wb = openpyxl.load_workbook(JOURNAL_FILE)
        ws = wb['Journal']
        journal_entries = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and len(row) >= 5:
                entry = {
                    'row_id': idx,
                    'tanggal': row[0],
                    'keterangan': row[1],
                    'akun': row[2],
                    'debit': row[3] if row[3] else 0,
                    'kredit': row[4] if row[4] else 0
                }
                journal_entries.append(entry)
                logger.debug(f"Loaded journal entry: {entry}")
        logger.info(f"Loaded {len(journal_entries)} journal entries for journal route")
    except FileNotFoundError:
        logger.warning("jurnal.xlsx file not found for journal route, using sample data.")
        journal_entries = [
            {'row_id': 2, 'tanggal': '2025-11-01', 'keterangan': 'Penjualan Madu', 'akun': 'Kas', 'debit': 1000000, 'kredit': 0},
            {'row_id': 3, 'tanggal': '2025-11-01', 'keterangan': 'Penjualan Madu', 'akun': 'Penjualan', 'debit': 0, 'kredit': 1000000},
        ]
    except Exception as e:
        logger.error(f"Error loading jurnal.xlsx for journal route: {e}")
        journal_entries = []

    messages = get_flashed_messages()
    return render_template('journal.html', journal_entries=journal_entries, messages=messages)

@app.route('/delete_journal/<int:row_id>', methods=['GET'])
@login_required
def delete_journal(row_id):
    try:
        jurnal_path = JOURNAL_FILE
        wb = openpyxl.load_workbook(jurnal_path)
        ws = wb['Journal']
        if row_id <= 1 or row_id > ws.max_row:
            return redirect(url_for('journal'))

        # Before deleting, check if this journal entry affects stock
        # Identify item and quantity from the journal entry row to add back stock
        row = ws[row_id]
        if row and len(row) >= 5:
            tanggal = row[0].value
            keterangan = row[1].value
            akun = row[2].value
            debit = row[3].value if row[3].value else 0
            kredit = row[4].value if row[4].value else 0

            if keterangan and 'penjualan' in str(keterangan).lower():
                # Try to find product name in keterangan or akun
                product_name_found = None
                inventory_data = load_inventory()
                inventory_names = [item['name'] for item in inventory_data]
                for name in inventory_names:
                    if name.lower() in keterangan.lower():
                        product_name_found = name
                        break
                if product_name_found:
                    qty_to_increase = 0
                    # Determine qty based on debit or kredit fields (reverse of sale)
                    # Assumption: kredit field has amount for sales
                    if kredit and kredit > 0:
                        qty_to_increase = int(kredit / next((item['selling_price'] for item in inventory_data if item['name'] == product_name_found), 1))
                    elif debit and debit > 0:
                        qty_to_increase = int(debit / next((item['selling_price'] for item in inventory_data if item['name'] == product_name_found), 1))
                    if qty_to_increase > 0:
                        update_inventory_stock(product_name_found, qty_to_increase)

        # Delete the journal row
        ws.delete_rows(row_id)
        wb.save(jurnal_path)
        logger.info(f"Deleted journal entry at row {row_id}")
    except Exception as e:
        logger.error(f"Error deleting journal entry at row {row_id}: {e}")
    return redirect(url_for('journal'))

def update_inventory_stock(item_name, qty_change):
    """
    Update the stock quantity of the item with item_name in the Inventory sheet
    by adding qty_change (positive to increase stock, negative to decrease stock).
    """
    try:
        inventory_path = INVENTORY_FILE
        wb = openpyxl.load_workbook(inventory_path)
        if 'Inventory' not in wb.sheetnames:
            logger.error("Inventory sheet not found in databasesia.xlsx")
            return False
        ws = wb['Inventory']
        item_found = False

        for row in range(2, ws.max_row + 1):  # Assuming first row is header
            cell_value = ws.cell(row=row, column=2).value  # Column 2: 'name'
            if cell_value and cell_value.strip().lower() == item_name.strip().lower():
                current_stock = ws.cell(row=row, column=3).value  # Column 3: 'stock'
                if current_stock is None:
                    current_stock = 0
                new_stock = int(current_stock) + qty_change
                if new_stock < 0:
                    new_stock = 0  # Prevent negative stock
                ws.cell(row=row, column=3, value=new_stock)
                item_found = True
                logger.info(f"Updated stock for '{item_name}': from {current_stock} to {new_stock}")
                break

        if item_found:
            wb.save(inventory_path)
            return True
        else:
            logger.warning(f"Item '{item_name}' not found in Inventory to update stock.")
            return False

    except Exception as e:
        logger.error(f"Error updating inventory stock: {e}")
        return False


from flask import redirect

@app.route('/input_transaksi', methods=['GET', 'POST'])
@login_required
def input_transaksi():
    accounts = [
        ('1-1100', 'Kas'),
        ('1-1200', 'Piutang usaha'),
        ('1-1300', 'Persediaan barang dagang'),
        ('1-1310', 'Persediaan stok madu gudang'),
        ('1-1400', 'Perlengkapan toko'),
        ('1-1500', 'Tanah'),
        ('1-1510', 'Bangunan'),
        ('1-1511', 'Akumulasi penyusutan bangunan'),
        ('1-1600', 'Kendaraan'),
        ('1-1610', 'Akumulasi penyusutan kendaraan'),
        ('1-1700', 'Peralatan'),
        ('1-1710', 'Akumulasi penyusutan peralatan'),
        ('2-2100', 'Hutang dagang'),
        ('3-3000', 'Modal'),
        ('4-4000', 'Penjualan barang dagang'),
        ('4-4100', 'Retur penjualan'),
        ('5-5000', 'Harga pokok penjualan'),
        ('6-6100', 'Beban telepon, air, dan listrik'),
        ('6-6200', 'Beban perlengkapan'),
        ('6-6300', 'Beban pemeliharaan'),
        ('6-6400', 'Beban gaji produksi'),
        ('6-6500', 'Beban gaji pemeliharaan lebah'),
        ('6-6600', 'Beban transportasi pemeliharaan lebah'),
        ('6-6700', 'Beban transportasi penjualan lebah'),
        ('6-6800', 'Beban depresiasi aktiva tetap'),
    ]
    akun_options = [f"{code} - {name}" for code, name in accounts]

    inventory_data = load_inventory()

    if request.method == 'POST':
        try:
            # Extract form data from POST
            jenis_transaksi = request.form.get('jenis_transaksi')
            tanggal = request.form.get('tanggal')
            keterangan = request.form.get('keterangan')

            debit_entries = []
            kredit_entries = []

            # Collect debit akun and amounts
            debit_index = 1
            while True:
                akun_key = f"akun_debit_{debit_index}"
                amount_key = f"debit_{debit_index}"
                if akun_key not in request.form:
                    break
                akun_val = request.form.get(akun_key)
                amount_val = request.form.get(amount_key)
                if akun_val and amount_val:
                    try:
                        amount_float = parse_amount(amount_val)
                        if amount_float > 0:
                            debit_entries.append({'akun': akun_val, 'amount': amount_float})
                    except ValueError:
                        pass
                debit_index += 1

            # Collect kredit akun and amounts
            kredit_index = 1
            while True:
                akun_key = f"akun_kredit_{kredit_index}"
                amount_key = f"kredit_{kredit_index}"
                if akun_key not in request.form:
                    break
                akun_val = request.form.get(akun_key)
                amount_val = request.form.get(amount_key)
                if akun_val and amount_val:
                    try:
                        amount_float = parse_amount(amount_val)
                        if amount_float > 0:
                            kredit_entries.append({'akun': akun_val, 'amount': amount_float})
                    except ValueError:
                        pass
                kredit_index += 1

            # Validation: there must be at least one debit and kredit entry
            if not debit_entries or not kredit_entries:
                error_msg = "Transaksi harus memiliki minimal satu akun debit dan satu akun kredit dengan jumlah > 0."
                return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

            # Validation: total debit must equal total kredit
            total_debit = sum(item['amount'] for item in debit_entries)
            total_kredit = sum(item['amount'] for item in kredit_entries)
            if abs(total_debit - total_kredit) > 0.01:  # small epsilon for float comparison
                error_msg = f"Total debit ({total_debit}) dan total kredit ({total_kredit}) harus sama."
                return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

            # Use absolute path for jurnal.xlsx
            jurnal_path = JOURNAL_FILE
            try:
                if not os.path.exists(jurnal_path):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = 'Journal'
                    ws.append(['Tanggal', 'Keterangan', 'Akun', 'Debit', 'Kredit'])
                    wb.save(jurnal_path)

                wb = openpyxl.load_workbook(jurnal_path)
                if 'Journal' in wb.sheetnames:
                    ws = wb['Journal']
                else:
                    ws = wb.create_sheet('Journal')
                    ws.append(['Tanggal', 'Keterangan', 'Akun', 'Debit', 'Kredit'])

                # Append debit entries (avoid duplicates)
                for entry in debit_entries:
                    row_data = [tanggal, keterangan, entry['akun'], entry['amount'], 0]
                    if journal_row_exists(ws, *row_data):
                        logger.info(f"Skipping duplicate debit journal row: {row_data}")
                        continue
                    ws.append(row_data)

                # Append kredit entries (avoid duplicates)
                for entry in kredit_entries:
                    row_data = [tanggal, keterangan, entry['akun'], 0, entry['amount']]
                    if journal_row_exists(ws, *row_data):
                        logger.info(f"Skipping duplicate kredit journal row: {row_data}")
                        continue
                    ws.append(row_data)

                logger.info(f"Journal entries prepared for saving to {jurnal_path}")

            except Exception as e:
                logger.error(f"Error saving journal entries: {e}")
                error_msg = f"Terjadi kesalahan saat menyimpan transaksi: {str(e)}"
                return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

            # Update stock based on explicit sales rows (Penjualan)
            # Also create automatic journal entries for COGS (Harga Pokok Penjualan)
            if jenis_transaksi == 'Penjualan':
                logger.debug(f"Processing Penjualan stock updates for keterangan: {keterangan}")

                sales_items = []
                index = 1
                while True:
                    product_key = f"product_{index}"
                    qty_key = f"quantity_{index}"
                    if product_key not in request.form:
                        break
                    product_code = request.form.get(product_key)
                    qty_val = request.form.get(qty_key)
                    index += 1

                    if not product_code or not qty_val:
                        continue

                    try:
                        qty = int(float(qty_val))
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid quantity value for {product_key}: {qty_val}")
                        continue

                    if qty <= 0:
                        continue

                    item = next((item for item in inventory_data if item['item_code'] == product_code), None)
                    if not item:
                        logger.warning(f"Product code {product_code} not found in inventory for sales stock update")
                        continue

                    if qty > safe_int(item.get('stock', 0)):
                        error_msg = f"Stok untuk {item['name']} tidak mencukupi. Stok tersedia: {item['stock']}"
                        return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

                    sales_items.append({
                        'product_code': product_code,
                        'product_name': item['name'],
                        'qty': qty,
                        'cost_price': safe_float(item.get('cost_price', 0)),
                        'selling_price': safe_float(item.get('selling_price', 0))
                    })

                if not sales_items:
                    error_msg = "Penjualan harus memiliki minimal satu produk."
                    return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

                for sale in sales_items:
                    cogs_amount = sale['qty'] * sale['cost_price']
                    auto_keterangan = f"{keterangan} - {sale['product_name']} [AUTO]"

                    debit_row = [tanggal, auto_keterangan, '5-5000 - Harga pokok penjualan', cogs_amount, 0]
                    credit_row = [tanggal, auto_keterangan, '1-1300 - Persediaan barang dagang', 0, cogs_amount]

                    if journal_row_exists(ws, *debit_row) or journal_row_exists(ws, *credit_row):
                        logger.info(f"Auto journal entries already exist for {auto_keterangan}, skipping stock update.")
                        continue

                    ws.append(debit_row)
                    ws.append(credit_row)

                    success = update_inventory_stock(sale['product_name'], -sale['qty'])
                    if success:
                        logger.info(f"Stock updated (Penjualan): {sale['product_name']} decreased by {sale['qty']}")
                    else:
                        logger.error(f"Failed to update stock (Penjualan) for: {sale['product_name']}")

                wb.save(jurnal_path)

            # Update stock based on explicit purchase rows (Pembelian)
            elif jenis_transaksi == 'Pembelian':
                logger.debug(f"Processing Pembelian stock updates for keterangan: {keterangan}")
                index = 1
                while True:
                    product_key = f"purchase_product_{index}"
                    qty_key = f"purchase_quantity_{index}"
                    if product_key not in request.form:
                        break
                    product_code = request.form.get(product_key)
                    qty_val = request.form.get(qty_key)
                    index += 1

                    if not product_code or not qty_val:
                        continue

                    try:
                        qty = int(float(qty_val))
                    except (ValueError, TypeError):
                        logger.warning(f"Invalid purchase quantity value for {product_key}: {qty_val}")
                        continue

                    if qty <= 0:
                        continue

                    item = next((item for item in inventory_data if item['item_code'] == product_code), None)
                    if not item:
                        logger.warning(f"Product code {product_code} not found in inventory for purchase stock update")
                        continue

                    product_name = item['name']
                    success = update_inventory_stock(product_name, qty)
                    if success:
                        logger.info(f"Stock updated (Pembelian): {product_name} increased by {qty}")
                    else:
                        logger.error(f"Failed to update stock (Pembelian) for: {product_name}")

                wb.save(jurnal_path)

            else:
                wb.save(jurnal_path)

            # Redirect to journal page after successful save
            flash("Transaksi berhasil disimpan.")
            return redirect(url_for('journal'))

        except Exception as e:
            error_msg = f"Terjadi kesalahan saat menyimpan transaksi: {str(e)}"
            return render_template('input_transaksi.html', akun_options=akun_options, inventory_data=inventory_data, error=error_msg)

    return render_template('input_transaksi.html', 
                           akun_options=akun_options, 
                           inventory_data=inventory_data)

def create_dummy_daftarsaldo():
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        bee_the_one_dir = DATA_DIR
        file_path = SALDO_FILE

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "daftar saldo awal"

        ws.append(['No Akun', 'Nama Akun', 'Saldo'])

        dummy_akun = [
            ['101', 'Kas', 10000000],
            ['102', 'Piutang Usaha', 5000000],
            ['103', 'Persediaan', 15000000],
            ['104', 'Peralatan', 20000000],
            ['201', 'Utang Usaha', 8000000],
            ['301', 'Modal', 42000000],
            ['401', 'Pendapatan Penjualan', 0],
            ['501', 'Beban Pokok Penjualan', 0],
            ['502', 'Beban Operasional', 0],
            ['503', 'Beban Depresiasi Aktiva Tetap', 0],
            ['504', 'Beban Lain-lain', 0],
        ]

        for akun in dummy_akun:
            ws.append(akun)

        os.makedirs(bee_the_one_dir, exist_ok=True)
        wb.save(file_path)
        logger.info(f"Dummy daftarsaldo.xlsx created at: {file_path}")
        return True

    except Exception as e:
        logger.error(f"Error creating dummy daftarsaldo: {e}")
        return False

@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))

@app.route('/menu_madu')
@login_required
def menu_madu():
    harga_jual_list = [
        105000, 125000, 120000, 100000, 125000, 120000,
        115000, 115000, 115000, 135000, 300000, 162500
    ]
    inventory_data = load_inventory()
    jenis_madu_list = [item['name'] for item in inventory_data[:len(harga_jual_list)]]

    menu_items = []
    for i in range(len(harga_jual_list)):
        no_item = i + 1
        jenis_madu = jenis_madu_list[i] if i < len(jenis_madu_list) else f'Jenis Madu {no_item}'
        harga_jual = harga_jual_list[i]
        menu_items.append({
            'no_item': no_item,
            'jenis_madu': jenis_madu,
            'harga_jual': harga_jual
        })

    return render_template('menu_madu.html', menu_items=menu_items)

@app.route('/saldo_awal', methods=['GET'])
@login_required
def saldo_awal():
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')
    saldo_data = []
    total_debit = 0
    total_kredit = 0
    try:
        # If requested period is in the future (e.g. Desember while today is still November),
        # do not show any Saldo Awal yet.
        if _is_future_period(tahun, bulan):
            return render_template('saldo_awal.html', saldo_data=saldo_data, tahun=tahun, bulan=bulan)

        # If requested period is before the first journal month, also show nothing.
        min_year, min_month = _get_min_journal_period()
        if min_year is not None and min_month is not None:
            month_code = MONTH_NAME_TO_NUM.get(bulan, None)
            if month_code:
                selected_year = int(tahun)
                selected_month = int(month_code)
                if (selected_year, selected_month) < (min_year, min_month):
                    return render_template('saldo_awal.html', saldo_data=saldo_data, tahun=tahun, bulan=bulan)

        file_path = SALDO_FILE
        wb = openpyxl.load_workbook(file_path)
        ws = wb['daftar saldo awal'] if 'daftar saldo awal' in wb.sheetnames else wb.active
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                continue
            if len(row) >= 2 and row[0] and row[1]:
                no_akun = str(row[0]).strip()
                nama_akun = str(row[1]).strip()
                if (no_akun.startswith('=') or 'total' in no_akun.lower() or 
                    'sum' in no_akun.lower() or no_akun == ''):
                    continue
                
                side = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                debit_amount = 0
                kredit_amount = 0
                
                if len(row) > 3 and row[3] is not None and row[3] != '':
                    try:
                        debit_amount = float(row[3])
                    except (ValueError, TypeError):
                        debit_amount = 0
                if len(row) > 4 and row[4] is not None and row[4] != '':
                    try:
                        kredit_amount = float(row[4])
                    except (ValueError, TypeError):
                        kredit_amount = 0
                        
                final_side = side
                if not final_side:
                    if debit_amount > 0:
                        final_side = 'Debit'
                    elif kredit_amount > 0:
                        final_side = 'Kredit'
                    else:
                        final_side = 'Debit'
                        
                saldo_data.append({
                    'no_akun': no_akun,
                    'nama_akun': nama_akun,
                    'side': final_side,
                    'debit': format_rupiah(debit_amount),
                    'kredit': format_rupiah(kredit_amount)
                })
                total_debit += debit_amount
                total_kredit += kredit_amount
    except Exception as e:
        return render_template('saldo_awal.html', saldo_data=saldo_data, tahun=tahun, bulan=bulan, error=f"Error memuat data: {str(e)}")
    
    if saldo_data:
        saldo_data.append({
            'no_akun': 'Total',
            'nama_akun': '',
            'side': '',
            'debit': format_rupiah(total_debit),
            'kredit': format_rupiah(total_kredit),
            'is_total': True
        })
        
    return render_template('saldo_awal.html', saldo_data=saldo_data, tahun=tahun, bulan=bulan)

@app.route('/test_inventory')
def test_inventory():
    """Route untuk test inventory data"""
    inventory_data = load_inventory()
    return {
        'count': len(inventory_data),
        'data': inventory_data
    }

def _parse_account_code_name(akun_raw):
    akun_str = str(akun_raw).strip()
    if ' - ' in akun_str:
        code, name = akun_str.split(' - ', 1)
        return code.strip(), name.strip()
    return akun_str, akun_str


def _load_opening_balances():
    opening = {}
    if not os.path.exists(SALDO_FILE):
        logger.warning(f"Opening balance file not found: {SALDO_FILE}")
        return opening
    try:
        wb = openpyxl.load_workbook(SALDO_FILE)
        ws = wb['daftar saldo awal'] if 'daftar saldo awal' in wb.sheetnames else wb.active

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or all(cell is None for cell in row):
                logger.debug(f"Skipping empty or None row {idx}")
                continue
            if len(row) >= 2 and row[0] and row[1]:
                no_akun = str(row[0]).strip()
                nama_akun = str(row[1]).strip()
                if (no_akun.startswith('=') or 'total' in no_akun.lower() or
                    'sum' in no_akun.lower() or no_akun == ''):
                    logger.debug(f"Skipping row {idx} due to no_akun filter: {no_akun}")
                    continue

                debit_amount = 0.0
                kredit_amount = 0.0

                if len(row) > 3 and row[3] is not None and row[3] != '':
                    try:
                        debit_amount = float(row[3])
                    except (ValueError, TypeError) as err:
                        logger.error(f"Row {idx} debit_amount conversion error: {err}, value: {row[3]}")
                        debit_amount = 0.0
                if len(row) > 4 and row[4] is not None and row[4] != '':
                    try:
                        kredit_amount = float(row[4])
                    except (ValueError, TypeError) as err:
                        logger.error(f"Row {idx} kredit_amount conversion error: {err}, value: {row[4]}")
                        kredit_amount = 0.0

                if no_akun not in opening:
                    opening[no_akun] = {
                        'no_akun': no_akun,
                        'nama_akun': nama_akun,
                        'debit': 0.0,
                        'kredit': 0.0,
                    }
                opening[no_akun]['debit'] += debit_amount
                opening[no_akun]['kredit'] += kredit_amount
            else:
                logger.debug(f"Row {idx} skipped due to insufficient length or missing values")
    except Exception as e:
        logger.error(f"Error loading opening balances from {SALDO_FILE}: {e}")

    return opening


def load_journal_entries(tahun=None, bulan=None):
    entries = []
    if not os.path.exists(JOURNAL_FILE):
        logger.warning(f"Journal file not found: {JOURNAL_FILE}")
        return entries
    try:
        wb = openpyxl.load_workbook(JOURNAL_FILE)
        if 'Journal' not in wb.sheetnames:
            logger.warning("'Journal' sheet not found in jurnal.xlsx")
            return entries
        ws = wb['Journal']

        month_num = None
        if bulan:
            month_num = MONTH_NAME_TO_NUM.get(bulan, None)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 3:
                continue
            raw_date = row[0]
            keterangan = row[1]
            akun_raw = row[2]
            if not akun_raw:
                continue

            date_obj = None
            if isinstance(raw_date, datetime):
                date_obj = raw_date.date()
            elif raw_date:
                try:
                    date_obj = datetime.strptime(str(raw_date).split(' ')[0], '%Y-%m-%d').date()
                except ValueError:
                    logger.warning(f"Unable to parse date in journal row {idx}: {raw_date}")

            if tahun and bulan and date_obj is not None and month_num is not None:
                if str(date_obj.year) != str(tahun) or f"{date_obj.month:02d}" != month_num:
                    continue

            no_akun, nama_akun = _parse_account_code_name(akun_raw)

            debit_val = row[3] if len(row) > 3 else 0
            kredit_val = row[4] if len(row) > 4 else 0
            try:
                debit = float(debit_val) if debit_val else 0.0
            except (ValueError, TypeError):
                debit = 0.0
            try:
                kredit = float(kredit_val) if kredit_val else 0.0
            except (ValueError, TypeError):
                kredit = 0.0

            entries.append({
                'row_index': idx,
                'tanggal': date_obj,
                'keterangan': keterangan,
                'no_akun': no_akun,
                'nama_akun': nama_akun,
                'debit': debit,
                'kredit': kredit,
            })
    except Exception as e:
        logger.error(f"Error loading journal entries from {JOURNAL_FILE}: {e}")

    return entries


def load_neraca_saldo_data(tahun=None, bulan=None):
    opening = _load_opening_balances()
    journal_entries = load_journal_entries(tahun, bulan)

    saldo_per_akun = {}

    for no_akun, acc in opening.items():
        saldo_per_akun[no_akun] = {
            'no_akun': no_akun,
            'nama_akun': acc['nama_akun'],
            'debit': acc['debit'],
            'kredit': acc['kredit'],
        }

    for entry in journal_entries:
        no_akun = entry['no_akun']
        nama_akun = entry['nama_akun']
        debit = entry['debit']
        kredit = entry['kredit']
        if no_akun not in saldo_per_akun:
            saldo_per_akun[no_akun] = {
                'no_akun': no_akun,
                'nama_akun': nama_akun,
                'debit': 0.0,
                'kredit': 0.0,
            }
        saldo_per_akun[no_akun]['debit'] += debit
        saldo_per_akun[no_akun]['kredit'] += kredit

    saldo_data = []
    total_debit = 0.0
    total_kredit = 0.0

    for acc in saldo_per_akun.values():
        debit_amount = acc['debit'] or 0.0
        kredit_amount = acc['kredit'] or 0.0
        if debit_amount == 0 and kredit_amount == 0:
            continue
        if debit_amount > kredit_amount:
            side = 'Debit'
        elif kredit_amount > debit_amount:
            side = 'Kredit'
        else:
            side = 'Debit'

        saldo_data.append({
            'no_akun': acc['no_akun'],
            'nama_akun': acc['nama_akun'],
            'side': side,
            'debit': debit_amount,
            'kredit': kredit_amount,
        })
        total_debit += debit_amount
        total_kredit += kredit_amount

    logger.debug(f"Total accounts in trial balance: {len(saldo_data)}, Total debit: {total_debit}, Total kredit: {total_kredit}")

    return saldo_data


def _get_min_journal_period():
    """Return (year, month) of earliest journal entry, or (None, None) if no entries.

    Used to hide Saldo Awal / Neraca Saldo for months before any activity exists.
    """
    entries = load_journal_entries()
    dates = [e['tanggal'] for e in entries if e.get('tanggal') is not None]
    if not dates:
        return None, None
    min_date = min(dates)
    return min_date.year, min_date.month


@app.route('/buku_besar')
@login_required
def buku_besar():
    search_query = request.args.get('search', '').strip().lower()
    tahun = request.args.get('tahun')
    bulan = request.args.get('bulan')

    opening_balances = _load_opening_balances()
    journal_entries = load_journal_entries(tahun, bulan)

    ledger_map = {}

    # Create ledger entries for all accounts in chart of accounts
    # This ensures accounts like "Penjualan" appear even with zero balance
    for no_akun, acc in opening_balances.items():
        nama_akun = acc['nama_akun']
        if search_query and search_query not in no_akun.lower() and search_query not in nama_akun.lower():
            continue
        saldo_awal = (acc['debit'] or 0.0) - (acc['kredit'] or 0.0)
        ledger_map[no_akun] = {
            'no_akun': no_akun,
            'nama_akun': nama_akun,
            'entries': [],
            'saldo_running': saldo_awal,
        }
        # Always show opening balance entry, even if zero (for revenue/expense accounts)
        if saldo_awal != 0 or no_akun.startswith('4') or no_akun.startswith('5') or no_akun.startswith('6'):
            ledger_map[no_akun]['entries'].append({
                'no': 1,
                'tanggal': '-',
                'keterangan': 'Saldo Awal',
                'debet': acc['debit'] or 0.0,
                'kredit': acc['kredit'] or 0.0,
                'saldo': saldo_awal,
            })

    sorted_journal = sorted(
        journal_entries,
        key=lambda e: (
            e['tanggal'] if e['tanggal'] is not None else datetime.min.date(),
            e['row_index'],
        ),
    )

    for entry in sorted_journal:
        no_akun = entry['no_akun']
        nama_akun = entry['nama_akun']
        if search_query and search_query not in no_akun.lower() and search_query not in nama_akun.lower():
            continue

        if no_akun not in ledger_map:
            ledger_map[no_akun] = {
                'no_akun': no_akun,
                'nama_akun': nama_akun,
                'entries': [],
                'saldo_running': 0.0,
            }

        ledger = ledger_map[no_akun]
        debit = entry['debit'] or 0.0
        kredit = entry['kredit'] or 0.0
        ledger['saldo_running'] += debit - kredit

        tanggal_str = ''
        if entry['tanggal'] is not None:
            try:
                tanggal_str = entry['tanggal'].strftime('%Y-%m-%d')
            except Exception:
                tanggal_str = str(entry['tanggal'])

        ledger['entries'].append({
            'no': len(ledger['entries']) + 1,
            'tanggal': tanggal_str or '-',
            'keterangan': entry['keterangan'] or '',
            'debet': debit,
            'kredit': kredit,
            'saldo': ledger['saldo_running'],
        })

    ledgers = list(ledger_map.values())
    ledgers.sort(key=lambda x: x['no_akun'])

    return render_template('buku_besar.html', ledgers=ledgers, search_query=search_query, tahun=tahun, bulan=bulan)

@app.route('/financial_reports')
@login_required
def financial_reports():
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')

    # Do not show financial reports for future periods
    if _is_future_period(tahun, bulan):
        return render_template('financial_reports.html',
                               saldo_data=[],
                               total_debit=format_rupiah(0),
                               total_kredit=format_rupiah(0),
                               tahun=tahun,
                               bulan=bulan)

    # Also hide periods before the first journal month
    min_year, min_month = _get_min_journal_period()
    if min_year is not None and min_month is not None:
        month_code = MONTH_NAME_TO_NUM.get(bulan, None)
        if month_code:
            selected_year = int(tahun)
            selected_month = int(month_code)
            if (selected_year, selected_month) < (min_year, min_month):
                return render_template('financial_reports.html',
                                       saldo_data=[],
                                       total_debit=format_rupiah(0),
                                       total_kredit=format_rupiah(0),
                                       tahun=tahun,
                                       bulan=bulan)

    saldo_data = load_neraca_saldo_data(tahun, bulan)
    total_debit = sum(item['debit'] for item in saldo_data)
    total_kredit = sum(item['kredit'] for item in saldo_data)

    # Format debit and kredit in saldo_data for display
    saldo_data_fmt = []
    for item in saldo_data:
        saldo_data_fmt.append({
            'no_akun': item['no_akun'],
            'nama_akun': item['nama_akun'],
            'side': item['side'],
            'debit': format_rupiah(item['debit']),
            'kredit': format_rupiah(item['kredit']),
        })

    return render_template('financial_reports.html',
                           saldo_data=saldo_data_fmt,
                           total_debit=format_rupiah(total_debit),
                           total_kredit=format_rupiah(total_kredit),
                           tahun=tahun,
                           bulan=bulan)

from flask import flash

@app.route('/jurnal_penutup', methods=['GET', 'POST'])
@login_required
def jurnal_penutup():
    saldo_closing_accounts = []
    closing_entries = []
    error = None
    message = None

    def load_closing_balances():
        saldo_data = []
        file_path = SALDO_FILE
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb['daftar saldo awal'] if 'daftar saldo awal' in wb.sheetnames else wb.active

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue
                if len(row) >= 2 and row[0] and row[1]:
                    no_akun = str(row[0]).strip()
                    nama_akun = str(row[1]).strip()
                    if (no_akun.startswith('=') or 'total' in no_akun.lower() or
                        'sum' in no_akun.lower() or no_akun == ''):
                        continue

                    side = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                    debit_amount = 0
                    kredit_amount = 0

                    if len(row) > 3 and row[3] is not None and row[3] != '':
                        try:
                            debit_amount = float(row[3])
                        except (ValueError, TypeError):
                            debit_amount = 0
                    if len(row) > 4 and row[4] is not None and row[4] != '':
                        try:
                            kredit_amount = float(row[4])
                        except (ValueError, TypeError):
                            kredit_amount = 0

                    final_side = side
                    if not final_side:
                        if debit_amount > 0:
                            final_side = 'Debit'
                        elif kredit_amount > 0:
                            final_side = 'Kredit'
                        else:
                            final_side = 'Debit'

                    # We consider only Pendapatan (income) and Beban (expenses) account types
                    # From the account number prefix, e.g., 4xxxx for Pendapatan, 5xxxx or 6xxxx for Beban
                    if no_akun.startswith('4') or no_akun.startswith('5') or no_akun.startswith('6'):
                        saldo_data.append({
                            'no_akun': no_akun,
                            'nama_akun': nama_akun,
                            'side': final_side,
                            'debit': debit_amount,
                            'kredit': kredit_amount
                        })
            return saldo_data
        except Exception as e:
            # nonlocal error
            error = f"Error loading saldo data: {str(e)}"
            return []

    if request.method == 'POST':
        saldo_closing_accounts = load_closing_balances()
        if not saldo_closing_accounts:
            error = error or "No closing accounts found to create jurnal penutup."
        else:
            try:
                jurnal_path = JOURNAL_FILE
                if not os.path.exists(jurnal_path):
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = 'Journal'
                    ws.append(['Tanggal', 'Keterangan', 'Akun', 'Debit', 'Kredit'])
                    wb.save(jurnal_path)

                wb = openpyxl.load_workbook(jurnal_path)
                if 'Journal' in wb.sheetnames:
                    ws = wb['Journal']
                else:
                    ws = wb.create_sheet('Journal')
                    ws.append(['Tanggal', 'Keterangan', 'Akun', 'Debit', 'Kredit'])

                today_str = pd.Timestamp.today().strftime('%Y-%m-%d')
                closing_entries = []
                for account in saldo_closing_accounts:
                    no_akun = account['no_akun']
                    nama_akun = account['nama_akun']
                    debit = account['debit']
                    kredit = account['kredit']
                    saldo = debit - kredit

                    if saldo == 0:
                        continue

                    akun_penutup = ''
                    debit_entry = 0
                    kredit_entry = 0

                    # Logic: Pendapatan (Income) accounts (4xxx) saldo normal kredit,
                    # so their saldo is kredit > debit -> nilai saldo positif artinya kredit,
                    # harus didebitkan ke akun penutup.
                    # Beban (Expense) accounts (5xxx,6xxx) saldo normal debit,
                    # jadi saldo > 0 artinya debit harus dikreditkan ke akun penutup.
                    if no_akun.startswith('4'):
                        akun_penutup = '3101 - Ikhtisar Laba Rugi'  # contoh akun penutup laba rugi
                        if saldo > 0:
                            debit_entry = saldo
                        else:
                            kredit_entry = abs(saldo)
                    elif no_akun.startswith('5') or no_akun.startswith('6'):
                        akun_penutup = '3101 - Ikhtisar Laba Rugi'
                        if saldo > 0:
                            kredit_entry = saldo
                        else:
                            debit_entry = abs(saldo)

                    if debit_entry > 0:
                        ws.append([today_str, f'Penutupan akun {no_akun} {nama_akun}', no_akun, debit_entry, 0])
                        ws.append([today_str, f'Penutupan ke akun penutup', akun_penutup, 0, debit_entry])
                        closing_entries.append({'akun': no_akun, 'debit': debit_entry, 'kredit': 0})
                        closing_entries.append({'akun': akun_penutup, 'debit': 0, 'kredit': debit_entry})
                    elif kredit_entry > 0:
                        ws.append([today_str, f'Penutupan akun {no_akun} {nama_akun}', no_akun, 0, kredit_entry])
                        ws.append([today_str, f'Penutupan ke akun penutup', akun_penutup, kredit_entry, 0])
                        closing_entries.append({'akun': no_akun, 'debit': 0, 'kredit': kredit_entry})
                        closing_entries.append({'akun': akun_penutup, 'debit': kredit_entry, 'kredit': 0})
                wb.save(jurnal_path)
                message = "Jurnal penutup berhasil dibuat."
            except Exception as e:
                error = f"Error creating jurnal penutup: {str(e)}"

    if request.method == 'GET' or saldo_closing_accounts == []:
        saldo_closing_accounts = load_closing_balances()

    return render_template('jurnal_penutup.html',
                           saldo_closing_accounts=saldo_closing_accounts,
                           closing_entries=closing_entries,
                           error=error,
                           message=message)

@app.route('/neraca_saldo')
@login_required
def neraca_saldo():
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')

    # Do not show Neraca Saldo for future periods
    if _is_future_period(tahun, bulan):
        return render_template('neraca_saldo.html',
                               saldo_data=[],
                               tahun=tahun,
                               bulan=bulan)

    # Also hide periods before the first journal month
    min_year, min_month = _get_min_journal_period()
    if min_year is not None and min_month is not None:
        month_code = MONTH_NAME_TO_NUM.get(bulan, None)
        if month_code:
            selected_year = int(tahun)
            selected_month = int(month_code)
            if (selected_year, selected_month) < (min_year, min_month):
                return render_template('neraca_saldo.html',
                                       saldo_data=[],
                                       tahun=tahun,
                                       bulan=bulan)

    saldo_data = load_neraca_saldo_data(tahun, bulan)

    # Format debit and kredit in saldo_data for display
    saldo_data_fmt = []
    total_debit = 0
    total_kredit = 0
    for item in saldo_data:
        debit = item.get('debit', 0) or 0
        kredit = item.get('kredit', 0) or 0
        total_debit += debit
        total_kredit += kredit
        saldo_data_fmt.append({
            'no_akun': item.get('no_akun', ''),
            'nama_akun': item.get('nama_akun', ''),
            'side': item.get('side', ''),
            'debit': format_rupiah(debit),
            'kredit': format_rupiah(kredit),
            'is_total': False,
        })

    if saldo_data_fmt:
        saldo_data_fmt.append({
            'no_akun': 'Total',
            'nama_akun': '',
            'side': '',
            'debit': format_rupiah(total_debit),
            'kredit': format_rupiah(total_kredit),
            'is_total': True,
        })

    return render_template('neraca_saldo.html',
                           saldo_data=saldo_data_fmt,
                           tahun=tahun,
                           bulan=bulan)


@app.route('/laba_rugi')
@login_required
def laba_rugi():
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')
    saldo_data = load_neraca_saldo_data(tahun, bulan)

    # Segregate accounts into categories based on account number prefix
    revenues = {}
    sales_returns = {}
    cogs = {}
    expenses = {}

    total_revenue = 0
    total_sales_returns = 0
    total_cogs = 0
    total_expenses = 0

    for item in saldo_data:
        no_akun = item.get('no_akun', '')
        nama_akun = item.get('nama_akun', '')
        debit = item.get('debit', 0) or 0
        kredit = item.get('kredit', 0) or 0
        saldo_normal = kredit - debit  # Revenues and sales returns normal balance Kredit
        saldo_debet = debit - kredit   # Expenses and COGS normal balance Debet

        if no_akun.startswith('4'):  # Pendapatan (Revenues)
            # Special case for Retur Penjualan (Sales Returns), account starting with 4-4100 or name containing 'Retur Penjualan'
            if 'retur' in nama_akun.lower() or 'retur penjualan' in nama_akun.lower() or no_akun.startswith('4-4100'):
                amount = abs(saldo_normal)
                sales_returns[nama_akun] = format_rupiah(amount)
                total_sales_returns += amount
            else:
                amount = saldo_normal
                revenues[nama_akun] = format_rupiah(amount if amount > 0 else 0)
                total_revenue += amount if amount > 0 else 0

        elif no_akun.startswith('5'):  # Harga Pokok Penjualan (COGS)
            amount = saldo_debet
            cogs[nama_akun] = format_rupiah(amount if amount > 0 else 0)
            total_cogs += amount if amount > 0 else 0

        elif no_akun.startswith('6'):  # Biaya (Expenses)
            amount = saldo_debet
            expenses[nama_akun] = format_rupiah(amount if amount > 0 else 0)
            total_expenses += amount if amount > 0 else 0

    net_sales_amount = total_revenue - total_sales_returns
    gross_profit_amount = net_sales_amount - total_cogs
    net_profit_amount = gross_profit_amount - total_expenses

    total_revenue_fmt = format_rupiah(total_revenue)
    total_sales_returns_fmt = format_rupiah(total_sales_returns)
    net_sales_fmt = format_rupiah(net_sales_amount)
    total_cogs_fmt = format_rupiah(total_cogs)
    gross_profit_fmt = format_rupiah(gross_profit_amount)
    total_expenses_fmt = format_rupiah(total_expenses)
    net_profit_fmt = format_rupiah(net_profit_amount)

    return render_template('laba_rugi.html',
                           revenues=revenues,
                           total_revenue=total_revenue_fmt,
                           sales_returns=sales_returns,
                           total_sales_returns=total_sales_returns_fmt,
                           net_sales=net_sales_fmt,
                           cogs=cogs,
                           total_cogs=total_cogs_fmt,
                           gross_profit=gross_profit_fmt,
                           expenses=expenses,
                           total_expenses=total_expenses_fmt,
                           net_profit=net_profit_fmt)

@app.route('/laporan_posisi_keuangan_detail')
@login_required
def laporan_posisi_keuangan_detail():
    tahun = request.args.get('tahun', '2025')
    bulan = request.args.get('bulan', 'November')
    saldo_data = load_neraca_saldo_data(tahun, bulan)

    # Classification of accounts into categories & subcategories for Balance Sheet (Posisi Keuangan)
    categories = {
        'Aktiva': {
            'name': 'AKTIVA',
            'subcategories': {
                'Aset Lancar': [],
                'Aset Tetap': [],
            }
        },
        'Kewajiban': {
            'name': 'KEWAJIBAN',
            'subcategories': {
                'Kewajiban': [],
            }
        },
        'Ekuitas': {
            'name': 'EKUITAS',
            'subcategories': {
                'Modal Awal': [],
                'Laba Bersih': [],
            }
        }
    }

    # Aggregate into top-level groups
    groups = [
        {
            'name': 'AKTIVA',
            'categories': [categories['Aktiva']],
            'total': ''
        },
        {
            'name': 'KEWAJIBAN DAN EKUITAS',
            'categories': [categories['Kewajiban'], categories['Ekuitas']],
            'total': ''
        }
    ]

    def classify_account(no_akun, nama_akun):
        aset_lancar_akun = ['101', '102', '103', '1310', '1400']
        aset_tetap_akun = ['104', '1500', '1510', '1511', '1600', '1610', '1700', '1710']
        kewajiban_akun = ['201']
        ekuitas_modal_akun = ['301']
        # For laba bersih, could match 'Laba Bersih' or derived later
        
        if no_akun in aset_lancar_akun or nama_akun.lower() in ['kas', 'piutang usaha', 'persediaan barang dagang', 'persediaan stok madu gudang', 'perlengkapan toko']:
            return 'Aktiva', 'Aset Lancar'
        elif no_akun in aset_tetap_akun or nama_akun.lower() in ['tanah', 'bangunan', 'akumulasi penyusutan bangunan', 'kendaraan', 'akumulasi penyusutan kendaraan', 'peralatan', 'akumulasi penyusutan peralatan']:
            return 'Aktiva', 'Aset Tetap'
        elif no_akun in kewajiban_akun or nama_akun.lower() == 'hutang dagang':
            return 'Kewajiban', 'Kewajiban'
        elif no_akun in ekuitas_modal_akun or nama_akun.lower() == 'modal awal':
            return 'Ekuitas', 'Modal Awal'
        elif 'laba bersih' in nama_akun.lower():
            # Assuming laba bersih is in equity but that might be summarized and not in neraca saldo
            return 'Ekuitas', 'Laba Bersih'
        else:
            # Skip other accounts (income/expense etc.)
            return None, None

    # Calculate modal_awal from neraca saldo accounts starting with '3'
    modal_awal = 0
    for item in saldo_data:
        if item.get('no_akun', '').startswith('3'):
            debit = item.get('debit', 0) or 0
            kredit = item.get('kredit', 0) or 0
            saldo = kredit - debit  # Assuming modal normal balance credit
            modal_awal += saldo

    # Calculate laba_bersih similar to laba_rugi()
    total_revenue = 0
    total_returns = 0
    total_cogs = 0
    total_expenses = 0

    for item in saldo_data:
        no_akun = item.get('no_akun', '')
        nama_akun = item.get('nama_akun', '')
        debit = item.get('debit', 0) or 0
        kredit = item.get('kredit', 0) or 0
        saldo_normal = kredit - debit  # Revenues and sales returns normal balance Kredit
        saldo_debet = debit - kredit   # Expenses and COGS normal balance Debet

        if no_akun.startswith('4'):  # Pendapatan (Revenues)
            if 'retur' in nama_akun.lower() or 'retur penjualan' in nama_akun.lower() or no_akun.startswith('4-4100'):
                amount = abs(saldo_normal)
                total_returns += amount
            else:
                amount = saldo_normal
                total_revenue += amount if amount > 0 else 0
        elif no_akun.startswith('5'):  # Harga Pokok Penjualan (COGS)
            amount = saldo_debet
            total_cogs += amount if amount > 0 else 0
        elif no_akun.startswith('6'):  # Biaya (Expenses)
            amount = saldo_debet
            total_expenses += amount if amount > 0 else 0

    laba_bersih = total_revenue - total_returns - total_cogs - total_expenses

    # Assign items to categories & subcategories
    for item in saldo_data:
        no_akun = item.get('no_akun', '')
        nama_akun = item.get('nama_akun', '')
        debit = item.get('debit', 0) or 0
        kredit = item.get('kredit', 0) or 0

        kategori, subkategori = classify_account(no_akun, nama_akun)

        if kategori is None or subkategori is None:
            logger.warning(f"Unclassified account in laporan_posisi_keuangan_detail: no_akun={no_akun}, nama_akun={nama_akun}")
            continue

        # For kewajiban (liabilities), saldo normal is kredit - debit, so invert calculation to have saldo adding when kredit > debit
        if kategori == 'Kewajiban' or (kategori == 'Ekuitas' and subkategori == 'Modal Awal'):
            saldo = kredit - debit
        else:
            saldo = debit - kredit

        # Negate saldo for accumulated depreciation accounts to show deduction
        if kategori == 'Aktiva' and subkategori == 'Aset Tetap' and 'akumulasi penyusutan' in nama_akun.lower():
            saldo = -abs(saldo)

        categories[kategori]['subcategories'][subkategori].append({
            'name': nama_akun,
            'amount': format_rupiah_for_report(abs(saldo)) if saldo >= 0 else '-' + format_rupiah_for_report(abs(saldo))
        })

    # Insert modal_awal and laba_bersih explicitly into Ekuitas subcategories
    categories['Ekuitas']['subcategories']['Modal Awal'] = [{
        'name': 'Modal Awal',
        'amount': format_rupiah_for_report(abs(modal_awal)) if modal_awal >= 0 else '-' + format_rupiah_for_report(abs(modal_awal))
    }]
    categories['Ekuitas']['subcategories']['Laba Bersih'] = [{
        'name': 'Laba Bersih',
        'amount': format_rupiah_for_report(abs(laba_bersih)) if laba_bersih >= 0 else '-' + format_rupiah_for_report(abs(laba_bersih))
    }]

    # Calculate totals for subcategories and categories
    total_aktiva = 0
    total_kewajiban = 0
    total_ekuitas = 0

    for kategori_key, kategori_val in categories.items():
        cat_total = 0
        subcategory_list = []
        for subcat_key, items in kategori_val['subcategories'].items():
            subcat_total = 0
            for item in items:
                # Remove "Rp " and dots to convert back to int for sum
                amount_str = item['amount'].replace('Rp ', '').replace('.', '').replace('-', '')
                amount_int = int(amount_str) if amount_str.isdigit() else 0

                # Subtract amount if it has '-' prefix (i.e. accumulated depreciation)
                if item['amount'].startswith('-'):
                    subcat_total -= amount_int
                else:
                    subcat_total += amount_int

            subcategory_list.append({
                'name': subcat_key,
                'item_list': items,
                'total': format_rupiah_for_report(subcat_total)
            })
            cat_total += subcat_total
        categories[kategori_key]['subcategories'] = subcategory_list
        categories[kategori_key]['total'] = format_rupiah_for_report(cat_total)
        if kategori_key == 'Aktiva':
            total_aktiva = cat_total
        elif kategori_key == 'Kewajiban':
            total_kewajiban = cat_total
        elif kategori_key == 'Ekuitas':
            total_ekuitas = cat_total

    total_kewajiban_dan_ekuitas = total_kewajiban + total_ekuitas

    # Aggregate into top-level groups
    groups = [
        {
            'name': 'AKTIVA',
            'categories': [categories['Aktiva']],
            'total': format_rupiah_for_report(total_aktiva)
        },
        {
            'name': 'KEWAJIBAN DAN EKUITAS',
            'categories': [categories['Kewajiban'], categories['Ekuitas']],
            'total': format_rupiah_for_report(total_kewajiban_dan_ekuitas)
        }
    ]

    financial_data = groups + [{
        'name': 'TOTALS',
        'total_aktiva': format_rupiah_for_report(total_aktiva),
        'total_kewajiban_dan_ekuitas': format_rupiah_for_report(total_kewajiban_dan_ekuitas)
    }]

    return render_template('laporan_posisi_keuangan_detail.html', financial_data=financial_data)

import openpyxl
import logging

# ...existing code...

# ...existing code...
       

@app.route('/laporan_perubahan_ekuitas')
@login_required
def laporan_perubahan_ekuitas():
    try:
        tahun = request.args.get('tahun', '2025')
        bulan = request.args.get('bulan', 'November')
        saldo_data = load_neraca_saldo_data(tahun, bulan)

        modal_awal = 0
        for item in saldo_data:
            no_akun = item.get('no_akun', '')
            if no_akun.startswith('3'):
                debit = item.get('debit', 0) or 0
                kredit = item.get('kredit', 0) or 0
                modal_awal += (kredit - debit)

        laba_bersih = 0
        total_revenue = 0
        total_returns = 0
        
        total_cogs = 0
        total_expenses = 0

        for item in saldo_data:
            no_akun = item.get('no_akun', '')
            nama_akun = item.get('nama_akun', '')
            debit = item.get('debit', 0) or 0
            kredit = item.get('kredit', 0) or 0

            saldo_normal = kredit - debit
            saldo_debet = debit - kredit

            if no_akun.startswith('4'):
                if 'retur' in nama_akun.lower():
                    total_returns += abs(saldo_normal)
                else:
                    total_revenue += max(saldo_normal, 0)
            elif no_akun.startswith('5'):
                total_cogs += max(saldo_debet, 0)
            elif no_akun.startswith('6'):
                total_expenses += max(saldo_debet, 0)

        laba_bersih = total_revenue - total_returns - total_cogs - total_expenses
        modal_akhir = modal_awal + laba_bersih
        
        report_data = [
            {'keterangan': 'Modal Awal', 'nominal': modal_awal},
            {'keterangan': 'Laba Bersih Tahun Berjalan', 'nominal': laba_bersih},
        ]
        
        return render_template('perubahan_ekuitas.html',
                             report_data=report_data,
                             modal_akhir=modal_akhir)
    
    except Exception as e:
        logger.error(f"Error di laporan_perubahan_ekuitas: {str(e)}")
        print(f"Error: {str(e)}")
        return render_template('perubahan_ekuitas.html',
                             report_data=[],
                             modal_akhir=0)

# ...existing code...

if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # Create database tables if they do not exist
    app.run(debug=True)